"""Export exam data to Excel (.xlsx)."""
import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.answer import Answer
from app.models.class_ import Class, Student
from app.models.exam import Exam, Question
from app.models.grading import Grading


async def export_answers_xlsx(db: AsyncSession, exam_id: int) -> bytes:
    """Export all student answers for an exam to Excel."""
    exam = await db.get(Exam, exam_id, options=[selectinload(Exam.questions)])
    if exam is None:
        raise ValueError(f"Exam {exam_id} not found")

    questions = sorted(exam.questions, key=lambda q: q.order_index)

    # Fetch all classes/students/answers
    classes_result = await db.execute(
        select(Class)
        .where(Class.exam_id == exam_id)
        .options(selectinload(Class.students).selectinload(Student.answers))
    )
    classes = classes_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "답안"

    # Header
    header = ["반", "학번", "이름"] + [f"문{q.number}" for q in questions]
    ws.append(header)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for cls in classes:
        for student in cls.students:
            answer_map = {a.question_id: a.answer_text for a in student.answers}
            row = [
                cls.name,
                student.student_number or "",
                student.name or "",
            ]
            for q in questions:
                row.append(answer_map.get(q.id, ""))
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_criteria_descriptions(matched_ids, rubric_json) -> str:
    """matched_criteria_ids(예: ['criteria_0', 'criteria_2'])를 description으로 변환."""
    if not matched_ids or not isinstance(rubric_json, dict):
        return ""
    criteria = rubric_json.get("criteria") or []
    if not criteria:
        return ""
    descriptions: list[str] = []
    for cid in matched_ids:
        try:
            idx = int(str(cid).split("_")[-1])
        except ValueError:
            continue
        if 0 <= idx < len(criteria):
            desc = (criteria[idx] or {}).get("description") or ""
            if desc:
                descriptions.append(desc)
    return " / ".join(descriptions)


async def export_gradings_xlsx(
    db: AsyncSession,
    exam_id: int,
    include_score: bool = True,
    include_rationale: bool = True,
    include_answer: bool = False,
    include_model_answer: bool = False,
    include_criteria: bool = False,
    include_total: bool = True,
) -> bytes:
    """Export all gradings for an exam to Excel.

    각 옵션은 컬럼 단위로 켜고 끌 수 있다. 기본정보(반/학번/이름)는 항상 포함.
    """
    exam = await db.get(Exam, exam_id, options=[selectinload(Exam.questions)])
    if exam is None:
        raise ValueError(f"Exam {exam_id} not found")

    questions = sorted(exam.questions, key=lambda q: q.order_index)

    classes_result = await db.execute(
        select(Class)
        .where(Class.exam_id == exam_id)
        .options(
            selectinload(Class.students)
            .selectinload(Student.answers)
            .selectinload(Answer.grading)
        )
    )
    classes = classes_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "채점결과"

    # 컬럼 구성 — 문항별 (학생답안, 점수, 채점이유, 모범답안, 매칭기준)
    per_question_cols: list[tuple[str, str]] = []  # (label_template, key)
    if include_answer:
        per_question_cols.append(("문{n} 학생답안", "answer"))
    if include_score:
        per_question_cols.append(("문{n}({max}점)", "score"))
    if include_rationale:
        per_question_cols.append(("문{n} 채점이유", "rationale"))
    if include_model_answer:
        per_question_cols.append(("문{n} 모범답안", "model_answer"))
    if include_criteria:
        per_question_cols.append(("문{n} 매칭기준", "criteria"))

    header: list[str] = ["반", "학번", "이름"]
    wide_col_indices: list[int] = []  # 자동 줄바꿈 할 컬럼 인덱스 (1-based)
    score_col_indices: list[int] = []  # 점수 컬럼 위치 (너비 좁게)

    col_idx = 4  # 기본정보 3개 다음
    for q in questions:
        for tpl, key in per_question_cols:
            header.append(tpl.format(n=q.number, max=q.max_score))
            if key in ("rationale", "answer", "model_answer", "criteria"):
                wide_col_indices.append(col_idx)
            elif key == "score":
                score_col_indices.append(col_idx)
            col_idx += 1
    if include_total:
        header.append("합계")
        total_col_idx = col_idx
    else:
        total_col_idx = None

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 문항별 rubric 미리 캐싱 (lazy loading 방지)
    rubric_by_qid: dict[int, dict] = {q.id: (q.rubric_json or {}) for q in questions}

    # 데이터 행
    for cls in classes:
        for student in cls.students:
            answer_map: dict[int, str] = {}
            score_map: dict[int, float] = {}
            rationale_map: dict[int, str] = {}
            criteria_map: dict[int, str] = {}
            for answer in student.answers:
                answer_map[answer.question_id] = answer.answer_text or ""
                g = answer.grading
                if g is not None:
                    score_map[answer.question_id] = float(g.score)
                    rationale_map[answer.question_id] = g.rationale or ""
                    criteria_map[answer.question_id] = _resolve_criteria_descriptions(
                        g.matched_criteria_ids, rubric_by_qid.get(answer.question_id),
                    )

            row: list = [cls.name, student.student_number or "", student.name or ""]
            total = 0.0
            for q in questions:
                for tpl, key in per_question_cols:
                    if key == "answer":
                        row.append(answer_map.get(q.id, ""))
                    elif key == "score":
                        s = score_map.get(q.id, "")
                        row.append(s)
                        if isinstance(s, float):
                            total += s
                    elif key == "rationale":
                        row.append(rationale_map.get(q.id, ""))
                    elif key == "model_answer":
                        row.append(q.model_answer or "")
                    elif key == "criteria":
                        row.append(criteria_map.get(q.id, ""))
            if include_total:
                row.append(total)
            ws.append(row)

    # 컬럼 너비
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(1)].width = 10  # 반
    ws.column_dimensions[get_column_letter(2)].width = 10  # 학번
    ws.column_dimensions[get_column_letter(3)].width = 10  # 이름
    for c in score_col_indices:
        ws.column_dimensions[get_column_letter(c)].width = 10
    for c in wide_col_indices:
        ws.column_dimensions[get_column_letter(c)].width = 40
    if total_col_idx:
        ws.column_dimensions[get_column_letter(total_col_idx)].width = 10

    # 긴 텍스트 컬럼 자동 줄바꿈
    if wide_col_indices:
        wrap = Alignment(wrap_text=True, vertical="top")
        for row_idx in range(2, ws.max_row + 1):
            for c in wide_col_indices:
                ws.cell(row=row_idx, column=c).alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
